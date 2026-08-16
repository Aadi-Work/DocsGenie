"""
XLSX renderer.

Writes only what the validated plan says, and inherits every visual property
from the template:

  * a merged region is written at its anchor and left merged;
  * a cloned row copies the template row's full style array, height, merges
    and number formats, then receives data;
  * formulas, protected cells and static regions are untouched by construction
    (the gate has already removed them from the plan).
"""

from __future__ import annotations

from copy import copy
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from ..ir import DocType
from ..logging_config import get_logger
from ..mapper import FillInstruction, FillPlan
from .base import BaseRenderer, RenderResult, WriteRecord, clear_if_unresolved_placeholder, register
from .xlsx_formula_shift import shift_formulas_in_sheet

log = get_logger("renderers.xlsx")


class XlsxRenderer(BaseRenderer):
    doc_type = DocType.XLSX

    def render(self, template_path: str, output_path: str, plan: FillPlan,
              clear_unresolved: bool = True) -> RenderResult:
        out = self.prepare_output(template_path, output_path)
        keep_vba = template_path.lower().endswith(".xlsm")
        wb = load_workbook(out, data_only=False, keep_vba=keep_vba)
        result = RenderResult(output_path=out)

        # tables first: row insertion shifts addresses below them, so single
        # fields are resolved afterwards against the shifted sheet.
        tables = [i for i in plan.writable() if i.kind == "table"]
        fields = [i for i in plan.writable() if i.kind == "field"]

        shifts: Dict[str, List[Tuple[int, int]]] = {}
        for ins in sorted(tables, key=lambda i: -(i.target.get("header_row") or 0)):
            rec, shift = self._write_table(wb, ins)
            result.records.append(rec)
            if shift:
                result.rows_added += shift[1]
                shifts.setdefault(ins.target.get("sheet", ""), []).append(shift)

        for ins in fields:
            result.records.append(self._write_field(wb, ins, shifts))

        if clear_unresolved:
            cleared = self._clear_unresolved_placeholders(wb)
            if cleared:
                log.info("Cleared %d unresolved placeholder(s) that had no fillable value", cleared)

        wb.save(out)
        wb.close()
        return result

    def _clear_unresolved_placeholders(self, wb) -> int:
        """Any {{token}} still literally present after the plan ran had no
        evidence or no role match - it stays blank, not visible as raw
        template syntax, in the shipped file."""
        n = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and not cell.value.startswith("="):
                        new_val = clear_if_unresolved_placeholder(cell.value)
                        if new_val != cell.value:
                            cell.value = new_val or None
                            n += 1
        return n

    # ------------------------------------------------------------------
    def _write_field(self, wb, ins: FillInstruction,
                     shifts: Dict[str, List[Tuple[int, int]]]) -> WriteRecord:
        sheet = ins.target.get("sheet")
        cell_ref = ins.target.get("cell")
        rec = WriteRecord(ins.node_id, ins.role, f"{sheet}!{cell_ref}", ins.value)
        try:
            ws: Worksheet = wb[sheet]
            row = int(ins.target.get("row") or 0)
            col = int(ins.target.get("col") or 0)
            if not row or not col:
                col_letter = "".join(c for c in cell_ref if c.isalpha())
                row = int("".join(c for c in cell_ref if c.isdigit()))
                col = column_index_from_string(col_letter)
            row += self._row_shift(shifts.get(sheet, []), row)

            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                rec.status = "skipped"
                rec.message = "refused: target became a formula"
                return rec
            cell.value = ins.value
            rec.target = f"{sheet}!{cell.coordinate}"
            rec.rows_written = 1
        except Exception as exc:                     # never let one field abort the run
            rec.status = "failed"
            rec.message = f"{type(exc).__name__}: {exc}"
        return rec

    @staticmethod
    def _row_shift(shifts: List[Tuple[int, int]], row: int) -> int:
        return sum(delta for at, delta in shifts if at <= row)

    # ------------------------------------------------------------------
    def _write_table(self, wb, ins: FillInstruction) -> Tuple[WriteRecord, Optional[Tuple[int, int]]]:
        sheet = ins.target.get("sheet")
        rec = WriteRecord(ins.node_id, ins.role, f"{sheet}!{ins.target.get('range')}")
        shift: Optional[Tuple[int, int]] = None
        try:
            ws: Worksheet = wb[sheet]
            rows = ins.rows or []
            header_row = int(ins.target.get("header_row"))
            template_row = int(ins.target.get("template_row", header_row + 1))
            col_map = self._column_map(ins)
            if not col_map:
                rec.status = "skipped"
                rec.message = "no writable columns resolved"
                return rec, None

            available = self._blank_rows_available(ws, template_row, col_map.values(),
                                                    max_row=ins.target.get("last_row"))
            needed = len(rows)
            if needed > available:
                extra = needed - available
                if available >= 1:
                    # Insert *inside* the band, at its current last row, so any
                    # formula range ending there (e.g. =COUNTA(B24:B26)) extends
                    # to cover the new rows too - the same thing Excel itself
                    # does when a row is inserted within a referenced range,
                    # as opposed to merely appended after it.
                    insert_at = template_row + available - 1
                else:
                    insert_at = template_row
                ws.insert_rows(insert_at, extra)
                for offset in range(extra):
                    self._clone_row_style(ws, template_row, insert_at + offset,
                                          min(col_map.values()), max(col_map.values()))
                n_fixed = shift_formulas_in_sheet(ws, insert_at, extra,
                                                  skip_rows=range(insert_at, insert_at + extra))
                if n_fixed:
                    rec.message = f"adjusted {n_fixed} formula(s) to include the inserted rows"
                shift = (insert_at, extra)

            for i, row_data in enumerate(rows):
                target_row = template_row + i
                for field_name, value in row_data.items():
                    col = col_map.get(field_name)
                    if col is None:
                        continue
                    cell = ws.cell(row=target_row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue                     # formula column: leave it to recalc
                    cell.value = value
            rec.rows_written = len(rows)
        except Exception as exc:
            rec.status = "failed"
            rec.message = f"{type(exc).__name__}: {exc}"
        return rec, shift

    def _column_map(self, ins: FillInstruction) -> Dict[str, int]:
        """field name -> column index, taken from the spec's location hints."""
        cols: Dict[str, int] = {}
        hints = ins.target.get("columns") or {}
        for fname, hint in hints.items():
            idx = hint.get("col_index")
            letter = hint.get("column")
            if idx:
                cols[fname] = int(idx)
            elif letter:
                cols[fname] = column_index_from_string(letter)
        return cols

    def _blank_rows_available(self, ws: Worksheet, start_row: int, columns,
                              max_row: Optional[int] = None) -> int:
        """
        Blank rows already sitting inside the template's own row band. Capped
        at the band's originally-detected last row, so a blank spacer row that
        merely follows the band (e.g. before a totals row) is never mistaken
        for writable table space - insertion is used for anything beyond it.
        """
        cols = list(columns)
        n, r = 0, start_row
        cap = max_row if max_row else start_row + 1000
        while r <= cap:
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in cols):
                break
            n += 1
            r += 1
        return max(n, 1)          # template_row itself always counts as writable

    def _clone_row_style(self, ws: Worksheet, src_row: int, dst_row: int,
                         min_col: int, max_col: int) -> None:
        """Style, height and merges cloned from the template row - nothing recreated."""
        for c in range(min_col, max_col + 1):
            src = ws.cell(row=src_row, column=c)
            dst = ws.cell(row=dst_row, column=c)
            if src.has_style:
                dst._style = copy(src._style)
            dst.number_format = src.number_format
        src_dim = ws.row_dimensions.get(src_row)
        if src_dim is not None and src_dim.height is not None:
            ws.row_dimensions[dst_row].height = src_dim.height

        for rng in list(ws.merged_cells.ranges):
            min_c, min_r, max_c, max_r = range_boundaries(str(rng))
            if min_r == src_row and max_r == src_row and min_c >= min_col and max_c <= max_col:
                ws.merge_cells(start_row=dst_row, start_column=min_c,
                               end_row=dst_row, end_column=max_c)


register("xlsx", XlsxRenderer)
register("xlsm", XlsxRenderer)
register("xltx", XlsxRenderer)
