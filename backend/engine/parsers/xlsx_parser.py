"""
XLSX -> Document Graph -> Template IR.

The sheet is never treated as a blob. It is scanned into regions:

    Workbook > Sheet > Section > {Label, ValueRegion, Table, Static}

Detection is purely structural (merges, borders, fills, number formats,
emptiness, adjacency). No template names, no hardcoded addresses, no
"if label == 'Meeting Date'". The semantic layer runs afterwards on the
graph this produces.
"""

from __future__ import annotations

from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from ..ir import (ColumnIR, DocType, Location, Node, NodeType, Section,
                  StyleSignals, TableIR, TemplateIR, ValueFormat)
from .base import (BLANK_LINE_RE, TemplateParser, detect_value_format,
                   find_placeholder, looks_like_label, register)

MAX_SCAN_ROWS = 2000
MAX_SCAN_COLS = 80


class XlsxParser(TemplateParser):
    doc_type = DocType.XLSX

    # ------------------------------------------------------------------
    def parse(self, path: str) -> TemplateIR:
        wb = load_workbook(path, data_only=False)          # formulas as written
        try:
            wb_vals = load_workbook(path, data_only=True)  # cached results, if any
        except Exception:
            wb_vals = None

        ir = TemplateIR(doc_type=DocType.XLSX, source_path=path,
                        meta={"sheets": wb.sheetnames,
                              "has_defined_names": bool(getattr(wb, "defined_names", None))})

        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            ws_vals = wb_vals[ws.title] if wb_vals is not None else None
            self._parse_sheet(ir, ws, ws_vals)

        wb.close()
        if wb_vals is not None:
            wb_vals.close()
        return ir

    # ------------------------------------------------------------------
    def _parse_sheet(self, ir: TemplateIR, ws: Worksheet, ws_vals) -> None:
        max_row = min(ws.max_row or 1, MAX_SCAN_ROWS)
        max_col = min(ws.max_column or 1, MAX_SCAN_COLS)
        if max_row < 1 or max_col < 1:
            return

        merges = self._merge_map(ws)
        sheet_protected = bool(getattr(ws.protection, "sheet", False))

        grid = self._read_grid(ws, ws_vals, max_row, max_col, merges, sheet_protected)

        sec = Section(section_id=f"sheet::{ws.title}", title=ws.title,
                      location=Location(DocType.XLSX, {"sheet": ws.title}))
        ir.sections.append(sec)

        consumed: Set[Tuple[int, int]] = set()
        # A label whose neighbour is a placeholder cell (not blank, but not
        # "occupied" either - it's waiting to be filled) registers its text
        # here so the placeholder gets classified using the human-written
        # label ("Attendees(YMSLI)") instead of just its raw token name
        # ("attendees_ymsli") when the loop reaches that cell.
        pending_label: Dict[Tuple[int, int], str] = {}

        # 1) section banners give context to every region beneath them
        banners = self._detect_section_banners(grid, max_row, max_col)

        # 2) tables next - they own whole row bands and would otherwise be
        #    misread as a pile of loose label/value pairs.
        tables = self._detect_tables(ws, grid, max_row, max_col, merges)
        for t in tables:
            t.section = self._nearest_banner(banners, t.header_row) or ws.title
            ir.tables.append(t)
            sec.table_ids.append(t.node_id)
            r0 = t.header_row
            r1 = t.template_row + max(t.existing_data_rows, 1)
            c0 = t.meta["min_col"]
            c1 = t.meta["max_col"]
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    consumed.add((r, c))

        # 3) label / value-region pairs, under the banner that governs them
        current_section = ws.title

        for r in range(1, max_row + 1):
            if r in banners:
                current_section = banners[r]["text"]
                cell = grid[(r, banners[r]["col"])]
                n = self._make_node(ws, cell, NodeType.SECTION_HEADER, current_section, ws.title)
                n.editable = False
                ir.nodes.append(n)
                sec.node_ids.append(n.node_id)
                continue

            for c in range(1, max_col + 1):
                if (r, c) in consumed:
                    continue
                cell = grid.get((r, c))
                if cell is None or cell["is_anchor"] is False:
                    continue

                text = cell["text"]
                ph = find_placeholder(text)

                # (a) explicit placeholder already in the template
                if ph:
                    node = self._make_node(ws, cell, NodeType.VALUE_REGION, text, current_section)
                    node.placeholder = ph
                    node.label = pending_label.get((r, c), ph.replace("_", " "))
                    node.is_empty = False
                    # Same "is this genuinely a multi-line block, or a plain
                    # single-line field" signal used for label+blank-cell
                    # pairs, computed from the cell's own formatting - a
                    # placeholder happening to be present shouldn't change
                    # how a cell's own merge span/wrap setting is read.
                    node.meta["multiline_capable"] = bool(
                        cell["merged"] and cell["merge_rows"] > 1) or cell["wrap"]
                    ir.nodes.append(node)
                    sec.node_ids.append(node.node_id)
                    consumed.add((r, c))
                    continue

                if not text.strip():
                    continue

                # a formula cell is output, never a label and never a target
                if cell["has_formula"]:
                    node = self._make_node(ws, cell, NodeType.STATIC, text, current_section)
                    node.editable = False
                    node.is_empty = False
                    ir.nodes.append(node)
                    sec.node_ids.append(node.node_id)
                    consumed.add((r, c))
                    continue

                # (b) label with an adjacent writable region
                if looks_like_label(text):
                    right = grid.get((r, c + 1))
                    below = grid.get((r + 1, c))
                    right_ph = find_placeholder(str(right["text"])) if right else None
                    below_ph = find_placeholder(str(below["text"])) if below else None

                    if right_ph or below_ph:
                        # The value for this label already exists as a
                        # placeholder cell right next to it (just not blank,
                        # so the ordinary blank-target search below would
                        # never find it and could wrongly latch onto some
                        # unrelated merged/bordered region instead). Register
                        # the richer label text for that cell and stop here -
                        # don't also search for a second target.
                        target_rc = (r, c + 1) if right_ph else (r + 1, c)
                        pending_label[target_rc] = text.strip().rstrip(":*").strip()
                        lab = self._make_node(ws, cell, NodeType.LABEL, text, current_section)
                        lab.editable = False
                        ir.nodes.append(lab)
                        sec.node_ids.append(lab.node_id)
                        consumed.add((r, c))
                        continue

                    target = self._find_value_region(grid, r, c, max_row, max_col, consumed)
                    if target is not None:
                        lab = self._make_node(ws, cell, NodeType.LABEL, text, current_section)
                        lab.editable = False
                        sample_text = str(target.get("text") or "")
                        val = self._make_node(
                            ws, target, NodeType.VALUE_REGION, sample_text, current_section
                        )
                        val.label = text.strip().rstrip(":*").strip()
                        val.label_node_id = lab.node_id
                        # Sample-filled templates: cell has demo content but is still a slot
                        if sample_text.strip():
                            val.is_empty = False
                            val.meta["sample_value"] = True
                        val.meta["multiline_capable"] = bool(
                            target["merged"] and target["merge_rows"] > 1) or target["wrap"]
                        ir.nodes.append(lab)
                        ir.nodes.append(val)
                        sec.node_ids.extend([lab.node_id, val.node_id])
                        consumed.add((r, c))
                        consumed.add((target["row"], target["col"]))
                        continue

                # (c) everything else with content is static - protect it
                node = self._make_node(ws, cell, NodeType.STATIC, text, current_section)
                node.editable = False
                node.is_empty = False
                ir.nodes.append(node)
                sec.node_ids.append(node.node_id)
                consumed.add((r, c))

        # 4) images / logos: pure branding, never writable
        for img in getattr(ws, "_images", []) or []:
            anchor = getattr(getattr(img, "anchor", None), "_from", None)
            loc = Location(DocType.XLSX, {"sheet": ws.title,
                                          "cell": f"{get_column_letter((anchor.col + 1) if anchor else 1)}"
                                                  f"{(anchor.row + 1) if anchor else 1}"})
            n = Node(node_id=f"{ws.title}::image::{len(ir.nodes)}", type=NodeType.IMAGE,
                     location=loc, editable=False, is_empty=False, section=ws.title)
            ir.nodes.append(n)

    # ------------------------------------------------------------------
    # grid construction
    # ------------------------------------------------------------------
    def _merge_map(self, ws: Worksheet) -> Dict[Tuple[int, int], dict]:
        m: Dict[Tuple[int, int], dict] = {}
        for rng in ws.merged_cells.ranges:
            min_c, min_r, max_c, max_r = range_boundaries(str(rng))
            info = {"range": str(rng), "anchor": (min_r, min_c),
                    "rows": max_r - min_r + 1, "cols": max_c - min_c + 1}
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    m[(r, c)] = info
        return m

    def _read_grid(self, ws, ws_vals, max_row, max_col, merges, sheet_protected) -> Dict[Tuple[int, int], dict]:
        grid: Dict[Tuple[int, int], dict] = {}
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                mi = merges.get((r, c))
                is_anchor = (mi is None) or (mi["anchor"] == (r, c))
                raw = cell.value
                has_formula = isinstance(raw, str) and raw.startswith("=")
                if has_formula and ws_vals is not None:
                    cached = ws_vals.cell(row=r, column=c).value
                else:
                    cached = raw
                text = "" if raw is None else str(raw)
                if BLANK_LINE_RE.match(text.strip()):
                    text = ""            # "______" is an input line, not content
                font = cell.font
                fill = cell.fill
                border = cell.border
                bordered = any(getattr(getattr(border, side, None), "style", None)
                               for side in ("left", "right", "top", "bottom"))
                filled = bool(fill and fill.fill_type not in (None, "none")
                              and getattr(fill.fgColor, "rgb", None) not in (None, "00000000", "FFFFFFFF"))
                grid[(r, c)] = {
                    "row": r, "col": c, "sheet": ws.title,
                    "coord": f"{get_column_letter(c)}{r}",
                    "text": text, "raw": raw, "cached": cached,
                    "has_formula": has_formula,
                    "is_anchor": is_anchor,
                    "merged": mi is not None,
                    "merge_range": mi["range"] if mi else None,
                    "merge_rows": mi["rows"] if mi else 1,
                    "merge_cols": mi["cols"] if mi else 1,
                    "bold": bool(font and font.bold),
                    "italic": bool(font and font.italic),
                    "underline": bool(font and font.underline),
                    "size": float(font.size) if font and font.size else None,
                    "font_name": font.name if font else None,
                    "filled": filled,
                    "bordered": bordered,
                    "number_format": cell.number_format,
                    "wrap": bool(cell.alignment and cell.alignment.wrap_text),
                    "align": cell.alignment.horizontal if cell.alignment else None,
                    "locked": bool(sheet_protected and cell.protection and cell.protection.locked),
                }
        return grid

    def _make_node(self, ws, cell: dict, ntype: NodeType, text: str, section: str) -> Node:
        loc = Location(DocType.XLSX, {
            "sheet": ws.title,
            "cell": cell["coord"],
            "range": cell["merge_range"] or cell["coord"],
            "row": cell["row"], "col": cell["col"],
        })
        style = StyleSignals(
            bold=cell["bold"], italic=cell["italic"], underline=cell["underline"],
            font_size=cell["size"], font_name=cell["font_name"],
            filled=cell["filled"], bordered=cell["bordered"], merged=cell["merged"],
            number_format=cell["number_format"], alignment=cell["align"],
            locked=cell["locked"],
        )
        node = Node(
            node_id=f"{ws.title}::{cell['coord']}",
            type=ntype, location=loc, text=text, section=section, style=style,
            value_format=detect_value_format(cell["number_format"], cell["text"]),
            editable=not (cell["has_formula"] or cell["locked"]),
            is_empty=not str(cell["text"]).strip(),
            has_formula=cell["has_formula"],
        )
        node.meta["merge_rows"] = cell["merge_rows"]
        node.meta["merge_cols"] = cell["merge_cols"]
        return node

    # ------------------------------------------------------------------
    # region detection
    # ------------------------------------------------------------------
    def _find_value_region(self, grid, r, c, max_row, max_col, consumed) -> Optional[dict]:
        """
        The writable region belonging to a label: right first (the dominant
        Western form-layout convention), then below.

        Accepts:
          1) empty merged/bordered/filled adjacent cells (classic blank form)
          2) sample-filled adjacent cells that are NOT themselves labels
             (common when the "template" is a filled example file)
        """
        # look right, allowing up to 2 spacer columns
        for dc in (1, 2, 3):
            cc = c + dc
            if cc > max_col:
                break
            cand = grid.get((r, cc))
            if cand is None:
                break
            if not cand["is_anchor"]:
                continue
            if (r, cc) in consumed or cand["has_formula"] or cand["locked"]:
                break
            text = str(cand["text"] or "").strip()
            if text:
                # Sample-as-template: treat non-label neighbour as replaceable value
                if dc == 1 and not looks_like_label(text) and not find_placeholder(text):
                    score = (2 if cand["merged"] else 0) + (1 if cand["bordered"] else 0) \
                            + (1 if cand["filled"] else 0) + 2  # adjacency
                    if score >= 2:
                        return cand
                break  # next label/value: stop
            score = (2 if cand["merged"] else 0) + (1 if cand["bordered"] else 0) \
                    + (1 if cand["filled"] else 0) + (2 if dc == 1 else 0)
            if score >= 2:
                return cand
        # look below (stacked layouts / big free-text boxes)
        below = grid.get((r + 1, c))
        if below and below["is_anchor"] and (r + 1, c) not in consumed \
                and not below["has_formula"] and not below["locked"]:
            btext = str(below["text"] or "").strip()
            if not btext and (below["merged"] or below["bordered"] or below["wrap"]):
                return below
            if btext and not looks_like_label(btext) and not find_placeholder(btext) \
                    and (below["merged"] or below["bordered"] or below["wrap"]):
                return below
        return None

    def _detect_section_banners(self, grid, max_row, max_col) -> Dict[int, dict]:
        """A row with exactly one emphasised text cell = a section title."""
        banners: Dict[int, dict] = {}
        for r in range(1, max_row + 1):
            filled_cells = [grid[(r, c)] for c in range(1, max_col + 1)
                            if grid.get((r, c)) and grid[(r, c)]["is_anchor"]
                            and str(grid[(r, c)]["text"]).strip()]
            if len(filled_cells) != 1:
                continue
            cell = filled_cells[0]
            emphasised = cell["bold"] or cell["filled"] or (cell["size"] or 0) >= 12
            # A banner spans or is shaded. Bold alone means "label", not "section" -
            # otherwise every 'Meeting Title' on its own row becomes a fake heading.
            spanning = cell["merge_cols"] > 1
            if emphasised and (spanning or cell["filled"]) and len(cell["text"]) <= 80:
                banners[r] = {"text": cell["text"].strip(), "col": cell["col"]}
        return banners

    @staticmethod
    def _nearest_banner(banners: Dict[int, dict], row: int) -> Optional[str]:
        above = [r for r in banners if r < row]
        return banners[max(above)]["text"] if above else None

    def _detect_tables(self, ws, grid, max_row, max_col, merges) -> List[TableIR]:
        """
        A header row is >=2 horizontally adjacent, emphasised, non-empty cells
        with a mostly-empty (or repeating) band underneath. The first row under
        the header becomes template_row - the row whose style gets cloned.
        """
        tables: List[TableIR] = []
        taken_rows: Set[int] = set()

        # openpyxl ListObjects are a gift when present - trust them first
        for tname, ref in (getattr(ws, "tables", {}) or {}).items():
            ref_str = ref if isinstance(ref, str) else getattr(ref, "ref", None)
            if not ref_str:
                continue
            min_c, min_r, max_c, max_r = range_boundaries(ref_str)
            t = self._build_table(ws, grid, min_r, min_c, max_c, max_r, name=tname)
            if t:
                tables.append(t)
                taken_rows.update(range(min_r, max_r + 1))

        for r in range(1, max_row):
            if r in taken_rows:
                continue
            runs = self._header_runs(grid, r, max_col)
            for (c0, c1) in runs:
                if (c1 - c0) < 1:
                    continue
                below_empty = self._band_is_writable(grid, r + 1, c0, c1)
                if not below_empty:
                    continue
                last = self._table_extent(grid, r + 1, c0, c1, max_row, max_col)
                t = self._build_table(ws, grid, r, c0, c1, last)
                if t:
                    tables.append(t)
                    taken_rows.update(range(r, last + 1))
        return tables

    def _header_runs(self, grid, r, max_col) -> List[Tuple[int, int]]:
        """
        A header cell names a column - it must never itself hold a
        placeholder (that's the fill target, one row down). A cell counts
        as a header if it's visually emphasised (bold/filled/bordered) OR if
        the cell directly beneath it holds a placeholder in the same column -
        the second signal is what lets a plain, unstyled template that relies
        purely on {{tokens}} still get its tables detected.
        """
        runs, start = [], None
        for c in range(1, max_col + 1):
            cell = grid.get((r, c))
            text = str(cell["text"]) if cell else ""
            is_placeholder_cell = bool(cell and find_placeholder(text))
            below = grid.get((r + 1, c))
            below_is_placeholder = bool(below and find_placeholder(str(below["text"])))
            ok = bool(cell and cell["is_anchor"] and text.strip()
                      and len(text) <= 60
                      and not cell["has_formula"] and not text.startswith("=")
                      and not is_placeholder_cell
                      and (cell["bold"] or cell["filled"] or cell["bordered"]
                          or below_is_placeholder))
            if ok and start is None:
                start = c
            elif not ok and start is not None:
                if c - 1 > start:
                    runs.append((start, c - 1))
                start = None
        if start is not None and max_col > start:
            runs.append((start, max_col))
        return runs

    def _band_is_writable(self, grid, r, c0, c1) -> bool:
        cells = [grid.get((r, c)) for c in range(c0, c1 + 1)]
        cells = [x for x in cells if x]
        if not cells:
            return False
        empties = sum(1 for x in cells if not str(x["text"]).strip())
        placeholders = sum(1 for x in cells if find_placeholder(str(x["text"])))
        formulas = sum(1 for x in cells if x["has_formula"])
        # A row directly beneath the header counts as "writable" whether its
        # cells are truly blank, or already hold {{placeholder}} tokens
        # waiting to be replaced - both are equally valid template rows.
        return (empties + placeholders) >= max(1, int(0.6 * len(cells))) and formulas < len(cells)

    def _table_extent(self, grid, first_data_row, c0, c1, max_row, max_col: Optional[int] = None) -> int:
        last = first_data_row
        r = first_data_row
        while r <= max_row:
            cells = [grid.get((r, c)) for c in range(c0, c1 + 1)]
            cells = [x for x in cells if x]
            if not cells:
                break
            structured = any(x["bordered"] or x["merged"]
                             or find_placeholder(str(x["text"])) for x in cells)
            has_text = any(str(x["text"]).strip() for x in cells)
            if not structured and not has_text:
                break
            # A row that itself looks like the start of a NEW table (its own
            # header run over a different column range) or a section banner
            # (one emphasised cell spanning the row) means the current
            # table's data band has ended here - even though the row is
            # still "structured" enough to otherwise look like more data.
            # Two adjacent tables with no blank row between them would
            # otherwise bleed into one, silently swallowing the second.
            if max_col is not None and r > first_data_row:
                other_runs = [run for run in self._header_runs(grid, r, max_col)
                             if run != (c0, c1)]
                if other_runs:
                    break
                filled = [grid[(r, c)] for c in range(1, max_col + 1)
                         if grid.get((r, c)) and grid[(r, c)]["is_anchor"]
                         and str(grid[(r, c)]["text"]).strip()]
                if len(filled) == 1 and (filled[0]["bold"] or filled[0]["filled"]):
                    break
            last = r
            r += 1
            if r - first_data_row > 500:
                break
        return last

    def _build_table(self, ws, grid, header_row, c0, c1, last_row, name: str = "") -> Optional[TableIR]:
        cols: List[ColumnIR] = []
        for i, c in enumerate(range(c0, c1 + 1)):
            cell = grid.get((header_row, c))
            if cell is None:
                continue
            probe = grid.get((header_row + 1, c)) or {}
            cols.append(ColumnIR(
                index=i,
                header_text=str(cell["text"]).strip(),
                location_hint={"column": get_column_letter(c), "col_index": c},
                value_format=detect_value_format(probe.get("number_format"), probe.get("text", "")),
                editable=not (probe.get("has_formula") or probe.get("locked")),
                has_formula=bool(probe.get("has_formula")),
            ))
        if len([c for c in cols if c.header_text]) < 2:
            return None

        template_row = header_row + 1
        existing = 0
        for r in range(template_row, last_row + 1):
            if any(str((grid.get((r, c)) or {}).get("text", "")).strip() for c in range(c0, c1 + 1)):
                existing += 1

        loc = Location(DocType.XLSX, {
            "sheet": ws.title,
            "range": f"{get_column_letter(c0)}{header_row}:{get_column_letter(c1)}{last_row}",
            "header_row": header_row, "template_row": template_row,
        })
        return TableIR(
            node_id=f"{ws.title}::table::{header_row}::{get_column_letter(c0)}",
            location=loc, columns=cols,
            header_row=header_row, template_row=template_row,
            existing_data_rows=existing,
            section=name or ws.title,
            meta={"min_col": c0, "max_col": c1, "last_row": last_row,
                  "list_object": bool(name)},
        )


register("xlsx", XlsxParser)
register("xlsm", XlsxParser)
register("xltx", XlsxParser)
